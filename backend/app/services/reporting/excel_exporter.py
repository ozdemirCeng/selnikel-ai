import io
import re
from typing import List
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class EngineeringExcelExporter:
    """Exports markdown calculation tables and data into beautifully formatted Excel spreadsheets."""

    @classmethod
    def generate_excel(cls, markdown_text: str, title: str = "Selnikel Mühendislik Verileri") -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mühendislik Verisi"

        # Styles
        header_fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Calibri", size=9, italic=True, color="64748B")
        data_font = Font(name="Calibri", size=10, color="1E293B")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # Title Block
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"SELNİKEL ENERJİ — {title.upper()}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical="center")

        ws.merge_cells("A2:E2")
        sub_cell = ws["A2"]
        sub_cell.value = "Otomatik Oluşturulan Mühendislik Tablosu & Hesap Verileri"
        sub_cell.font = sub_font
        sub_cell.alignment = Alignment(vertical="center")

        current_row = 4
        lines = markdown_text.split("\n")
        table_rows: List[List[str]] = []
        in_table = False

        for line in lines:
            trimmed = line.strip()

            # Markdown Table Row
            if trimmed.startswith("|") and trimmed.endswith("|"):
                if re.match(r"^\|[\s\-:|]+\|$", trimmed):
                    continue
                cols = [c.strip() for c in trimmed.strip("|").split("|")]
                table_rows.append(cols)
                in_table = True
                continue
            elif in_table and table_rows:
                current_row = cls._write_table_to_sheet(
                    ws, table_rows, current_row, header_fill, header_font, data_font, zebra_fill, thin_border
                )
                table_rows = []
                in_table = False
                current_row += 1

            # Headings outside tables
            if trimmed.startswith("#"):
                heading_text = trimmed.lstrip("#").strip()
                cell = ws.cell(row=current_row, column=1, value=heading_text)
                cell.font = Font(name="Calibri", size=12, bold=True, color="0284C7")
                current_row += 1
            elif trimmed and not trimmed.startswith("---"):
                cell = ws.cell(row=current_row, column=1, value=trimmed)
                cell.font = data_font
                current_row += 1

        if table_rows:
            current_row = cls._write_table_to_sheet(
                ws, table_rows, current_row, header_fill, header_font, data_font, zebra_fill, thin_border
            )

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if val:
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes

    @classmethod
    def _write_table_to_sheet(
        cls,
        ws,
        rows: List[List[str]],
        start_row: int,
        header_fill,
        header_font,
        data_font,
        zebra_fill,
        border,
    ) -> int:
        for r_idx, row_data in enumerate(rows):
            r = start_row + r_idx
            is_header = r_idx == 0
            for c_idx, cell_value in enumerate(row_data):
                c = c_idx + 1
                cell = ws.cell(row=r, column=c, value=cell_value)
                cell.border = border
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = data_font
                    cell.alignment = Alignment(vertical="center")
                    if r_idx % 2 == 0:
                        cell.fill = zebra_fill
        return start_row + len(rows) + 1
